"""
Document Ingestion Service.

Handles ingestion of various document formats:
- Excel (XLSX, XLS)
- Word (DOCX)
- PDF

Includes entity extraction for constituencies, candidates, parties, issues.
"""
from __future__ import annotations
from typing import Any, Dict, List, Optional, Tuple
from pathlib import Path
from dataclasses import dataclass
import hashlib
import json
import re
from datetime import datetime

from app.config import settings
from app.services.llm import get_llm
from app.models import EntityReference

# Conditional import for Lambda compatibility
try:
    from app.services.rag.local_store import LocalHybridIndex, DocumentChunk
    LOCAL_STORE_AVAILABLE = True
except ImportError:
    LOCAL_STORE_AVAILABLE = False
    LocalHybridIndex = None
    from dataclasses import dataclass as dc_dataclass, field as dc_field
    
    @dc_dataclass
    class DocumentChunk:
        doc_id: str
        chunk_id: str
        source_path: str
        text: str
        metadata: dict = dc_field(default_factory=dict)


@dataclass
class IngestResult:
    """Result of document ingestion."""
    doc_id: str
    file_name: str
    chunks_indexed: int
    entities_extracted: int
    metadata: Dict[str, Any]


class DocumentProcessor:
    """Process various document formats into text chunks."""
    
    def __init__(self, chunk_size: int = None, chunk_overlap: int = None):
        self.chunk_size = chunk_size or settings.chunk_size
        self.chunk_overlap = chunk_overlap or settings.chunk_overlap
    
    def process(self, file_path: Path) -> List[Dict[str, Any]]:
        """Process a file and return list of text chunks with metadata."""
        suffix = file_path.suffix.lower()
        
        if suffix in ['.xlsx', '.xls', '.xlsm']:
            return self._process_excel(file_path)
        elif suffix == '.docx':
            return self._process_docx(file_path)
        elif suffix == '.pdf':
            return self._process_pdf(file_path)
        elif suffix == '.txt':
            return self._process_text(file_path)
        elif suffix == '.csv':
            return self._process_csv(file_path)
        elif suffix == '.gz':
            # Handle gzipped CSV files
            if '.csv' in file_path.stem.lower():
                return self._process_csv_gz(file_path)
            else:
                raise ValueError(f"Unsupported gzipped file format: {suffix}")
        else:
            raise ValueError(f"Unsupported file format: {suffix}")
    
    def _process_excel(self, file_path: Path) -> List[Dict[str, Any]]:
        """Process Excel file."""
        import openpyxl
        
        chunks = []
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Get headers from first row
            headers = []
            rows = list(sheet.iter_rows(values_only=True))
            if rows:
                headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
            
            # Process data rows
            current_chunk = []
            for row_idx, row in enumerate(rows[1:], start=2):
                row_text = " | ".join([
                    f"{headers[i]}: {str(v)}" 
                    for i, v in enumerate(row) if v is not None
                ])
                
                if row_text.strip():
                    current_chunk.append(row_text)
                
                # Create chunk when we have enough content
                chunk_text = "\n".join(current_chunk)
                if len(chunk_text) >= self.chunk_size:
                    chunks.append({
                        "text": chunk_text,
                        "metadata": {
                            "sheet": sheet_name,
                            "start_row": row_idx - len(current_chunk) + 1,
                            "end_row": row_idx,
                            "source_type": "excel"
                        }
                    })
                    # Keep overlap
                    overlap_rows = max(1, int(len(current_chunk) * self.chunk_overlap / self.chunk_size))
                    current_chunk = current_chunk[-overlap_rows:]
            
            # Add remaining content
            if current_chunk:
                chunks.append({
                    "text": "\n".join(current_chunk),
                    "metadata": {
                        "sheet": sheet_name,
                        "source_type": "excel"
                    }
                })
        
        wb.close()
        return chunks
    
    def _process_docx(self, file_path: Path) -> List[Dict[str, Any]]:
        """Process Word document."""
        from docx import Document
        
        doc = Document(file_path)
        chunks = []
        current_chunk = []
        current_length = 0
        
        for para_idx, para in enumerate(doc.paragraphs):
            text = para.text.strip()
            if not text:
                continue
            
            current_chunk.append(text)
            current_length += len(text)
            
            if current_length >= self.chunk_size:
                chunks.append({
                    "text": "\n\n".join(current_chunk),
                    "metadata": {
                        "paragraph_start": para_idx - len(current_chunk) + 1,
                        "source_type": "docx"
                    }
                })
                # Keep overlap
                overlap_paras = max(1, int(len(current_chunk) * self.chunk_overlap / self.chunk_size))
                current_chunk = current_chunk[-overlap_paras:]
                current_length = sum(len(p) for p in current_chunk)
        
        # Add remaining content
        if current_chunk:
            chunks.append({
                "text": "\n\n".join(current_chunk),
                "metadata": {"source_type": "docx"}
            })
        
        return chunks
    
    def _process_pdf(self, file_path: Path) -> List[Dict[str, Any]]:
        """Process PDF document."""
        from pypdf import PdfReader
        
        reader = PdfReader(file_path)
        chunks = []
        current_chunk = []
        current_length = 0
        
        for page_idx, page in enumerate(reader.pages):
            text = page.extract_text()
            if not text or not text.strip():
                continue
            
            # Split page into paragraphs
            paragraphs = text.split('\n\n')
            
            for para in paragraphs:
                para = para.strip()
                if not para:
                    continue
                
                current_chunk.append(para)
                current_length += len(para)
                
                if current_length >= self.chunk_size:
                    chunks.append({
                        "text": "\n\n".join(current_chunk),
                        "metadata": {
                            "page": page_idx + 1,
                            "source_type": "pdf"
                        }
                    })
                    overlap_paras = max(1, int(len(current_chunk) * self.chunk_overlap / self.chunk_size))
                    current_chunk = current_chunk[-overlap_paras:]
                    current_length = sum(len(p) for p in current_chunk)
        
        # Add remaining content
        if current_chunk:
            chunks.append({
                "text": "\n\n".join(current_chunk),
                "metadata": {"source_type": "pdf"}
            })
        
        return chunks
    
    def _process_text(self, file_path: Path) -> List[Dict[str, Any]]:
        """Process plain text file."""
        text = file_path.read_text(encoding='utf-8', errors='ignore')

        # For chat-style TXT (e.g., WhatsApp exports), we want fewer, larger chunks
        # to reduce embedding calls and avoid long ingest times.
        # This does NOT increase token risk because we still cap by chunk_size.
        effective_chunk_size = max(self.chunk_size, 2000)

        # WhatsApp/chat exports often have single-newline lines, not double-newline paragraphs.
        # If we split only on "\n\n" we may end up with ONE gigantic "paragraph" which then
        # becomes ONE huge chunk and breaks embedding limits (8192 tokens / ~32k chars).
        if "\n\n" not in text and "\n" in text:
            paragraphs = text.splitlines()
        else:
            # Split into paragraphs
            paragraphs = text.split('\n\n')
        
        chunks = []
        current_chunk = []
        current_length = 0
        
        for para in paragraphs:
            para = para.strip()
            if not para:
                continue

            # If a single paragraph/line is extremely large, split it further to avoid
            # creating an oversized chunk that will fail embeddings.
            if len(para) > self.chunk_size * 2:
                # Flush current buffer first
                if current_chunk:
                    chunks.append({
                        "text": "\n\n".join(current_chunk),
                        "metadata": {"source_type": "txt"}
                    })
                    current_chunk = []
                    current_length = 0

                # Reuse existing helper to split by lines into smaller chunks
                for sub in self._split_large_chunk(para, group_name=file_path.stem):
                    # Ensure source_type is set for txt
                    sub.setdefault("metadata", {})
                    sub["metadata"]["source_type"] = "txt"
                    chunks.append(sub)
                continue
            
            current_chunk.append(para)
            current_length += len(para)
            
            if current_length >= effective_chunk_size:
                chunks.append({
                    "text": "\n\n".join(current_chunk),
                    "metadata": {"source_type": "txt"}
                })
                overlap_paras = max(1, int(len(current_chunk) * self.chunk_overlap / effective_chunk_size))
                current_chunk = current_chunk[-overlap_paras:]
                current_length = sum(len(p) for p in current_chunk)
        
        if current_chunk:
            chunks.append({
                "text": "\n\n".join(current_chunk),
                "metadata": {"source_type": "txt"}
            })
        
        return chunks
    
    def _process_csv(self, file_path: Path) -> List[Dict[str, Any]]:
        """Process CSV file with intelligent chunking for political data."""
        import pandas as pd
        
        try:
            # Try different encodings
            for encoding in ['utf-8', 'latin-1', 'cp1252']:
                try:
                    df = pd.read_csv(file_path, encoding=encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                df = pd.read_csv(file_path, encoding='utf-8', errors='ignore')
        except Exception as e:
            print(f"Warning: Error reading CSV {file_path}: {e}")
            return []
        
        chunks = []
        
        # Detect if this is electoral data based on columns
        columns_lower = [c.lower() for c in df.columns]
        
        is_electoral = any(col in ' '.join(columns_lower) for col in 
                          ['constituency', 'candidate', 'party', 'votes', 'winner', 'margin', 'ac_name', 'pc_name'])
        
        if is_electoral:
            # Group by constituency or create intelligent chunks
            group_cols = []
            for potential_col in ['Constituency_Name', 'AC_Name', 'PC Name', 'Constituency']:
                if potential_col in df.columns:
                    group_cols.append(potential_col)
                    break
            
            if group_cols:
                # Group by constituency and create detailed chunks
                for group_val, group_df in df.groupby(group_cols[0]):
                    chunk_text = self._dataframe_to_text(group_df, str(group_val))
                    
                    if len(chunk_text) > self.chunk_size * 2:
                        # Split large constituency data into multiple chunks
                        sub_chunks = self._split_large_chunk(chunk_text, str(group_val))
                        chunks.extend(sub_chunks)
                    else:
                        chunks.append({
                            "text": chunk_text,
                            "metadata": {
                                "source_type": "csv",
                                "constituency": str(group_val),
                                "data_type": "electoral"
                            }
                        })
            else:
                # No grouping column, chunk by rows
                chunks = self._chunk_dataframe_by_rows(df, "electoral")
        else:
            # Generic CSV - chunk by rows
            chunks = self._chunk_dataframe_by_rows(df, "tabular")
        
        return chunks
    
    def _process_csv_gz(self, file_path: Path) -> List[Dict[str, Any]]:
        """Process gzipped CSV file."""
        import gzip
        import pandas as pd
        
        try:
            with gzip.open(file_path, 'rt', encoding='utf-8') as f:
                df = pd.read_csv(f)
        except Exception as e:
            print(f"Warning: Error reading gzipped CSV {file_path}: {e}")
            return []
        
        # Use same logic as regular CSV
        chunks = []
        columns_lower = [c.lower() for c in df.columns]
        is_electoral = any(col in ' '.join(columns_lower) for col in 
                          ['constituency', 'candidate', 'party', 'votes', 'winner'])
        
        if is_electoral:
            group_cols = []
            for potential_col in ['Constituency_Name', 'AC_Name', 'PC Name', 'Constituency']:
                if potential_col in df.columns:
                    group_cols.append(potential_col)
                    break
            
            if group_cols:
                for group_val, group_df in df.groupby(group_cols[0]):
                    chunk_text = self._dataframe_to_text(group_df, str(group_val))
                    chunks.append({
                        "text": chunk_text,
                        "metadata": {
                            "source_type": "csv_gz",
                            "constituency": str(group_val),
                            "data_type": "electoral"
                        }
                    })
            else:
                chunks = self._chunk_dataframe_by_rows(df, "electoral")
        else:
            chunks = self._chunk_dataframe_by_rows(df, "tabular")
        
        return chunks
    
    def _dataframe_to_text(self, df, group_name: str) -> str:
        """Convert DataFrame to readable text for electoral data."""
        import pandas as pd
        
        lines = [f"=== {group_name} ===\n"]
        
        for _, row in df.iterrows():
            row_parts = []
            for col in df.columns:
                val = row[col]
                if pd.notna(val) and str(val).strip():
                    # Clean column name
                    col_clean = col.replace('_', ' ').title()
                    row_parts.append(f"{col_clean}: {val}")
            
            if row_parts:
                lines.append(" | ".join(row_parts))
        
        return "\n".join(lines)
    
    def _chunk_dataframe_by_rows(self, df, data_type: str) -> List[Dict[str, Any]]:
        """Chunk DataFrame by rows when no grouping is possible."""
        import pandas as pd
        
        chunks = []
        current_rows = []
        current_length = 0
        
        # Include header info in each chunk
        header_info = f"Columns: {', '.join(df.columns)}\n\n"
        
        for idx, row in df.iterrows():
            row_parts = []
            for col in df.columns:
                val = row[col]
                if pd.notna(val) and str(val).strip():
                    row_parts.append(f"{col}: {val}")
            
            row_text = " | ".join(row_parts)
            current_rows.append(row_text)
            current_length += len(row_text)
            
            if current_length >= self.chunk_size:
                chunk_text = header_info + "\n".join(current_rows)
                chunks.append({
                    "text": chunk_text,
                    "metadata": {
                        "source_type": "csv",
                        "data_type": data_type,
                        "row_count": len(current_rows)
                    }
                })
                # Keep some overlap
                overlap_count = max(2, len(current_rows) // 5)
                current_rows = current_rows[-overlap_count:]
                current_length = sum(len(r) for r in current_rows)
        
        # Add remaining rows
        if current_rows:
            chunk_text = header_info + "\n".join(current_rows)
            chunks.append({
                "text": chunk_text,
                "metadata": {
                    "source_type": "csv",
                    "data_type": data_type,
                    "row_count": len(current_rows)
                }
            })
        
        return chunks
    
    def _split_large_chunk(self, text: str, group_name: str) -> List[Dict[str, Any]]:
        """Split large text chunk into smaller pieces."""
        lines = text.split('\n')
        chunks = []
        current_lines = [lines[0]]  # Keep header
        current_length = len(lines[0])
        
        for line in lines[1:]:
            current_lines.append(line)
            current_length += len(line)
            
            if current_length >= self.chunk_size:
                chunks.append({
                    "text": "\n".join(current_lines),
                    "metadata": {
                        "source_type": "csv",
                        "constituency": group_name,
                        "data_type": "electoral"
                    }
                })
                current_lines = [lines[0]]  # Keep header for next chunk
                current_length = len(lines[0])
        
        if len(current_lines) > 1:
            chunks.append({
                "text": "\n".join(current_lines),
                "metadata": {
                    "source_type": "csv",
                    "constituency": group_name,
                    "data_type": "electoral"
                }
            })
        
        return chunks


class EntityExtractor:
    """Extract political entities from text using LLM."""
    
    def __init__(self):
        self.llm = get_llm()
        self.entity_types = settings.entity_types
    
    def extract(self, text: str, doc_id: str) -> List[EntityReference]:
        """Extract entities from text."""
        system = """You extract political entities from text. Output JSON only.
Entity types: constituency, candidate, party, issue, leader"""
        
        prompt = f"""Extract all political entities from this text.

Text:
{text[:2000]}

Return JSON array:
[
    {{"entity_type": "constituency", "entity_name": "Nandigram", "attributes": {{"district": "Purba Medinipur"}}}},
    {{"entity_type": "party", "entity_name": "BJP", "attributes": {{"full_name": "Bharatiya Janata Party"}}}},
    {{"entity_type": "candidate", "entity_name": "Suvendu Adhikari", "attributes": {{"party": "BJP"}}}},
    ...
]"""
        
        try:
            response = self.llm.generate(prompt, system=system, temperature=0.1)
            match = re.search(r'\[[\s\S]*\]', response.text)
            if match:
                entities_data = json.loads(match.group())
                
                entities = []
                for e in entities_data:
                    if e.get("entity_type") in self.entity_types:
                        entities.append(EntityReference(
                            entity_type=e["entity_type"],
                            entity_name=e["entity_name"],
                            attributes=e.get("attributes", {}),
                            source_doc_ids=[doc_id],
                            confidence=0.8
                        ))
                return entities
        except Exception:
            pass
        
        return []


def ingest_file(
    index: LocalHybridIndex,
    file_path: Path,
    document_id: str = None,
    extract_entities: bool = True
) -> Tuple[str, int, List[EntityReference]]:
    """
    Ingest a document file into the RAG index.
    
    Args:
        index: The hybrid index to add chunks to
        file_path: Path to the document
        document_id: Optional custom document ID
        extract_entities: Whether to extract entities
    
    Returns:
        Tuple of (document_id, chunks_indexed, entities_extracted)
    """
    # Generate document ID
    if not document_id:
        content_hash = hashlib.md5(file_path.read_bytes()).hexdigest()[:8]
        document_id = f"{file_path.stem}_{content_hash}"
    
    # Process document
    processor = DocumentProcessor()
    raw_chunks = processor.process(file_path)
    
    # Create DocumentChunk objects
    chunks = []
    for i, chunk_data in enumerate(raw_chunks):
        chunk_id = f"{document_id}_chunk_{i}"
        chunks.append(DocumentChunk(
            doc_id=document_id,
            chunk_id=chunk_id,
            source_path=str(file_path),
            text=chunk_data["text"],
            metadata={
                **chunk_data.get("metadata", {}),
                "file_name": file_path.name,
                "indexed_at": datetime.now().isoformat()
            }
        ))
    
    # Add to index
    index.add_chunks(chunks)
    
    # Extract entities if requested
    entities = []
    if extract_entities and chunks:
        extractor = EntityExtractor()
        # Sample a few chunks for entity extraction
        sample_text = "\n\n".join([c.text for c in chunks[:3]])
        entities = extractor.extract(sample_text, document_id)
    
    return document_id, len(chunks), entities
